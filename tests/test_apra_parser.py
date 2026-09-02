from __future__ import annotations

import zipfile
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

import ausecon_mcp.parsers.apra_xlsx as apra_parser
from ausecon_mcp.catalogue.apra import APRA_CATALOGUE
from ausecon_mcp.parsers.apra_xlsx import parse_apra_xlsx


def _xlsx_bytes(rows_by_sheet: dict[str, list[list[object]]]) -> bytes:
    workbook = Workbook()
    first = True
    for title, rows in rows_by_sheet.items():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = title
        for row in rows:
            sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _row_record_table_map() -> dict:
    return {
        "table_1": {
            "sheet": "Table 1",
            "layout": "row_records",
            "title": "Table 1",
            "unit": "$ million",
            "frequency": "Monthly",
            "header_row": 2,
            "data_start_row": 3,
            "date_column": 1,
            "dimension_columns": {"abn": 2, "institution": 3},
            "series_start_column": 4,
            "identity_columns": ["abn"],
        }
    }


def test_parse_apra_xlsx_rejects_oversized_compressed_payload(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(apra_parser, "MAX_APRA_XLSX_BYTES", 8, raising=False)

    with pytest.raises(ValueError, match="exceeds maximum APRA XLSX size"):
        parse_apra_xlsx(
            b"not-an-xlsx",
            publication_id="TEST_PUBLICATION",
            title="Test APRA publication",
            frequency="Monthly",
            table_maps=_row_record_table_map(),
        )


def test_parse_apra_xlsx_rejects_excessive_uncompressed_zip_payload(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("xl/worksheets/sheet1.xml", "x" * 32)
    monkeypatch.setattr(apra_parser, "MAX_APRA_XLSX_UNCOMPRESSED_BYTES", 16, raising=False)

    with pytest.raises(ValueError, match="uncompressed APRA XLSX payload"):
        parse_apra_xlsx(
            buffer.getvalue(),
            publication_id="TEST_PUBLICATION",
            title="Test APRA publication",
            frequency="Monthly",
            table_maps=_row_record_table_map(),
        )


def test_parse_apra_xlsx_rejects_excessive_table_dimensions(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = _xlsx_bytes(
        {
            "Table 1": [
                ["($million)"],
                ["Period", "ABN", "Institution Name", "Total residents assets"],
                [datetime(2024, 1, 31), 11111111111, "Example Bank", 100.5],
                [datetime(2024, 2, 29), 11111111111, "Example Bank", 110.0],
            ]
        }
    )
    monkeypatch.setattr(apra_parser, "MAX_APRA_TABLE_ROWS", 3, raising=False)

    with pytest.raises(ValueError, match="exceeds maximum APRA table rows"):
        parse_apra_xlsx(
            workbook,
            publication_id="TEST_PUBLICATION",
            title="Test APRA publication",
            frequency="Monthly",
            table_maps=_row_record_table_map(),
        )


def test_parse_apra_xlsx_normalises_row_record_tables() -> None:
    workbook = _xlsx_bytes(
        {
            "Table 1": [
                ["($million)"],
                [
                    "Period",
                    "ABN",
                    "Institution Name",
                    "Total residents assets",
                    "Deposits by households",
                ],
                [datetime(2024, 1, 31), 11111111111, "Example Bank", 100.5, "*"],
                [datetime(2024, 2, 29), 11111111111, "Example Bank", 110.0, 55.5],
                [None, None, None, None, None],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="TEST_PUBLICATION",
        title="Test APRA publication",
        frequency="Monthly",
        table_maps={
            "table_1": {
                "sheet": "Table 1",
                "layout": "row_records",
                "title": "Table 1",
                "unit": "$ million",
                "frequency": "Monthly",
                "header_row": 2,
                "data_start_row": 3,
                "date_column": 1,
                "dimension_columns": {
                    "abn": 2,
                    "institution": 3,
                },
                "series_start_column": 4,
                "identity_columns": ["abn"],
            }
        },
    )

    assert payload["metadata"]["source"] == "apra"
    assert payload["metadata"]["dataset_id"] == "TEST_PUBLICATION"
    assert payload["metadata"]["title"] == "Test APRA publication"
    assert payload["series"][0]["series_id"] == (
        "TEST_PUBLICATION:table_1:11111111111:total_residents_assets"
    )
    assert payload["series"][0]["label"] == "Example Bank - Total residents assets"
    assert payload["series"][0]["unit"] == "$ million"
    assert payload["observations"][0] == {
        "date": "2024-01-31",
        "series_id": "TEST_PUBLICATION:table_1:11111111111:total_residents_assets",
        "value": 100.5,
        "dimensions": {
            "table": {"code": "table_1", "label": "Table 1"},
            "abn": {"code": "11111111111", "label": "11111111111"},
            "institution": {"code": "Example Bank", "label": "Example Bank"},
        },
        "status": None,
        "comment": None,
    }
    assert payload["observations"][1]["raw_value"] == "*"
    assert payload["observations"][1]["value"] is None


def test_parse_apra_xlsx_normalises_matrix_tables_with_section_labels() -> None:
    workbook = _xlsx_bytes(
        {
            "Tab 1b": [
                ["Table 1b  Residential property exposures"],
                ["($ million, Level 2)"],
                [None, None, "Quarter end", None],
                [None, None, datetime(2024, 3, 31), datetime(2024, 6, 30)],
                [None, None, None, None],
                ["Credit outstanding", None, None, None],
                ["Total credit outstanding", None, 1000.0, 1050.5],
                ["Owner-occupied", None, 700.0, 725.0],
                [None, None, None, None],
                ["Memo row with no observations", None, None, None],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="TEST_PUBLICATION",
        title="Test APRA publication",
        frequency="Quarterly",
        table_maps={
            "tab_1b": {
                "sheet": "Tab 1b",
                "layout": "matrix",
                "title": "Residential property exposures",
                "unit": "$ million",
                "frequency": "Quarterly",
                "date_row": 4,
                "date_start_column": 3,
                "data_start_row": 6,
                "label_column": 1,
            }
        },
    )

    assert payload["series"][0]["series_id"] == (
        "TEST_PUBLICATION:tab_1b:credit_outstanding:total_credit_outstanding"
    )
    assert payload["series"][0]["label"] == "Credit outstanding: Total credit outstanding"
    assert payload["series"][0]["dimensions"]["section"]["label"] == "Credit outstanding"
    assert payload["observations"][0] == {
        "date": "2024-03-31",
        "series_id": "TEST_PUBLICATION:tab_1b:credit_outstanding:total_credit_outstanding",
        "value": 1000.0,
        "dimensions": {
            "table": {"code": "tab_1b", "label": "Residential property exposures"},
            "section": {"code": "credit_outstanding", "label": "Credit outstanding"},
        },
        "status": None,
        "comment": None,
    }
    target_observations = [
        item
        for item in payload["observations"]
        if item["series_id"]
        == "TEST_PUBLICATION:tab_1b:credit_outstanding:total_credit_outstanding"
    ]
    assert [item["date"] for item in target_observations] == [
        "2024-03-31",
        "2024-06-30",
    ]


def test_parse_apra_xlsx_normalises_period_row_tables() -> None:
    workbook = _xlsx_bytes(
        {
            "T1": [
                [None, "Coverage of hospital treatment tables", None, None, None, None],
                ["Quarter ended", None, None, None, "NSW", "Aust."],
                ["Dec", "2024", "Coverage ('000)", None, 3990.4, 12400.8],
                [None, None, "% Population", None, 0.46, 0.45],
                ["Mar", "2025", "Coverage ('000)", None, 4010.0, 12500.0],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="TEST_PUBLICATION",
        title="Test PHI membership publication",
        frequency="Quarterly",
        table_maps={
            "t1": {
                "sheet": "T1",
                "layout": "period_rows",
                "title": "Hospital treatment membership coverage",
                "frequency": "Quarterly",
                "header_row": 2,
                "data_start_row": 3,
                "month_column": 1,
                "year_column": 2,
                "metric_column": 3,
                "series_start_column": 5,
            }
        },
    )

    assert payload["series"][1]["series_id"] == "TEST_PUBLICATION:t1:aust:coverage_000"
    assert payload["series"][1]["label"] == "Aust. - Coverage ('000)"
    assert payload["observations"][1] == {
        "date": "2024-12-31",
        "series_id": "TEST_PUBLICATION:t1:aust:coverage_000",
        "value": 12400.8,
        "dimensions": {
            "table": {"code": "t1", "label": "Hospital treatment membership coverage"},
            "region": {"code": "aust", "label": "Aust."},
            "metric": {"code": "coverage_000", "label": "Coverage ('000)"},
        },
        "status": None,
        "comment": None,
    }


def test_parse_apra_xlsx_supports_super_industry_catalogue_fixture() -> None:
    workbook = _xlsx_bytes(
        {
            "Table 2": [
                [None],
                [None],
                [None],
                [None],
                [None, datetime(2024, 3, 31), datetime(2024, 6, 30)],
                ["Total RSE member assets", 100.0, 105.0],
                ["Total RSE member accounts", 12.0, 13.0],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="APRA_SUPER_INDUSTRY",
        title=APRA_CATALOGUE["APRA_SUPER_INDUSTRY"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_SUPER_INDUSTRY"]["tables"],
        table_id="table_2",
    )

    assert {series["series_id"] for series in payload["series"]} == {
        "APRA_SUPER_INDUSTRY:table_2:total_rse_member_assets",
        "APRA_SUPER_INDUSTRY:table_2:total_rse_member_accounts",
    }


def test_parse_apra_xlsx_supports_super_industry_table_5_member_accounts_fixture() -> None:
    """Regression coverage for Bug 3 (2026-08-02): superannuation_member_accounts was
    repointed to Table 5's "By fund type: Total industry" row, the genuine
    whole-of-industry member-account total (as opposed to Table 2's MySuper-only
    figures). This mirrors the real workbook's layout: a "By fund type" section
    header (blank-data row) followed by per-fund-type rows and a "Total industry"
    aggregate row.
    """
    workbook = _xlsx_bytes(
        {
            "Table 5": [
                [None],
                ["Total industry"],
                [None],
                [None, "('000)"],
                [None, datetime(2025, 12, 31), datetime(2026, 3, 31)],
                ["By fund type"],
                ["Corporate funds", 125.0, 123.0],
                ["Industry funds", 14669.0, 14669.0],
                ["Total industry", 24912.0, 25150.0],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="APRA_SUPER_INDUSTRY",
        title=APRA_CATALOGUE["APRA_SUPER_INDUSTRY"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_SUPER_INDUSTRY"]["tables"],
        table_id="table_5",
    )

    target = "APRA_SUPER_INDUSTRY:table_5:by_fund_type:total_industry"
    assert target in {series["series_id"] for series in payload["series"]}
    observations = {
        obs["date"]: obs["value"] for obs in payload["observations"] if obs["series_id"] == target
    }
    assert observations == {"2025-12-31": 24912.0, "2026-03-31": 25150.0}


def test_parse_apra_xlsx_supports_super_fund_level_catalogue_fixture() -> None:
    workbook = _xlsx_bytes(
        {
            "Table 1": [
                [None],
                [None],
                [None],
                [
                    "Period",
                    "Fund",
                    "ABN",
                    "Regulatory classification",
                    "Fund type",
                    "Membership base",
                    "Licensee",
                    "Ownership type",
                    "Profit status",
                    "Board structure",
                    "Member accounts",
                ],
                [None],
                [None],
                [
                    datetime(2024, 3, 31),
                    "Example Fund",
                    12345678901,
                    "RSE",
                    "Public offer",
                    "Open",
                    "Example Licensee",
                    "Retail",
                    "Profit-to-member",
                    "Board",
                    2500.0,
                ],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="APRA_SUPER_FUND_LEVEL",
        title=APRA_CATALOGUE["APRA_SUPER_FUND_LEVEL"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_SUPER_FUND_LEVEL"]["tables"],
        table_id="table_1",
    )

    assert payload["series"][0]["series_id"] == (
        "APRA_SUPER_FUND_LEVEL:table_1:12345678901:member_accounts"
    )


def test_parse_apra_xlsx_supports_general_insurance_catalogue_fixture() -> None:
    workbook = _xlsx_bytes(
        {
            "Database": [
                [
                    "Period",
                    "Data item",
                    "Category",
                    "Subject",
                    "Stock or flow",
                    "Industry segment",
                    "Industry segment group",
                    "Class of business",
                    "Class of business category",
                    "Class of business group",
                    "Counterparty grade",
                    "State or territory",
                    "Stress scenario type",
                    "Value",
                ],
                [
                    datetime(2024, 3, 31),
                    "Insurance revenue",
                    "Insurance revenue",
                    "Financial Performance",
                    "Flow",
                    "Total industry",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    100.0,
                ],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="APRA_GENERAL_INSURANCE_PERFORMANCE",
        title=APRA_CATALOGUE["APRA_GENERAL_INSURANCE_PERFORMANCE"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_GENERAL_INSURANCE_PERFORMANCE"]["tables"],
        table_id="database",
    )

    assert payload["series"][0]["series_id"] == (
        "APRA_GENERAL_INSURANCE_PERFORMANCE:database:"
        "insurance_revenue:insurance_revenue:financial_performance:flow:total_industry:value"
    )


def test_parse_apra_xlsx_supports_life_insurance_catalogue_fixture() -> None:
    workbook = _xlsx_bytes(
        {
            "Database": [
                [
                    "Period",
                    "Data item",
                    "Subject",
                    "Category",
                    "Stock or flow",
                    "Reporting structure",
                    "Product group",
                    "Product group type",
                    "Superannuation or ordinary business",
                    "Value",
                ],
                [
                    datetime(2024, 3, 31),
                    "Insurance revenue",
                    "Financial Performance",
                    "Insurance revenue",
                    "Flow",
                    "Total Entity",
                    "",
                    "",
                    "",
                    200.0,
                ],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="APRA_LIFE_INSURANCE_PERFORMANCE",
        title=APRA_CATALOGUE["APRA_LIFE_INSURANCE_PERFORMANCE"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_LIFE_INSURANCE_PERFORMANCE"]["tables"],
        table_id="database",
    )

    assert payload["series"][0]["series_id"] == (
        "APRA_LIFE_INSURANCE_PERFORMANCE:database:"
        "insurance_revenue:financial_performance:insurance_revenue:flow:total_entity:value"
    )


def test_parse_apra_xlsx_supports_private_health_insurance_catalogue_fixture() -> None:
    workbook = _xlsx_bytes(
        {
            "Database": [
                ["Period", "Data item", "Subject", "Category", "Stock or flow", "Value"],
                [
                    datetime(2024, 3, 31),
                    "HIB Premium Revenue",
                    "Financial Performance Supplementary",
                    "Revenue",
                    "Flow",
                    300.0,
                ],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="APRA_PHI_PERFORMANCE",
        title=APRA_CATALOGUE["APRA_PHI_PERFORMANCE"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_PHI_PERFORMANCE"]["tables"],
        table_id="database",
    )

    assert payload["series"][0]["series_id"] == (
        "APRA_PHI_PERFORMANCE:database:"
        "hib_premium_revenue:financial_performance_supplementary:revenue:flow:value"
    )


def test_parse_apra_xlsx_supports_phi_membership_catalogue_fixture() -> None:
    workbook = _xlsx_bytes(
        {
            "T1": [
                [None, "Coverage of hospital treatment tables", None, None, None, None],
                ["Quarter ended", None, None, None, "NSW", "Aust."],
                ["Dec", "2024", "Coverage ('000)", None, 3990.4, 12400.8],
            ]
        }
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="APRA_PHI_MEMBERSHIP",
        title=APRA_CATALOGUE["APRA_PHI_MEMBERSHIP"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_PHI_MEMBERSHIP"]["tables"],
        table_id="t1",
    )

    assert payload["series"][1]["series_id"] == "APRA_PHI_MEMBERSHIP:t1:aust:coverage_000"


def test_parse_apra_xlsx_can_limit_to_one_declared_table() -> None:
    workbook = _xlsx_bytes(
        {
            "A": [
                [None],
                ["Period", "Entity", "Metric"],
                [datetime(2024, 1, 31), "Entity A", 1.0],
            ],
            "B": [
                [None],
                ["Period", "Entity", "Metric"],
                [datetime(2024, 1, 31), "Entity B", 2.0],
            ],
        }
    )
    table_maps = {
        "a": {
            "sheet": "A",
            "layout": "row_records",
            "title": "A",
            "unit": "count",
            "frequency": "Monthly",
            "header_row": 2,
            "data_start_row": 3,
            "date_column": 1,
            "dimension_columns": {"entity": 2},
            "series_start_column": 3,
            "identity_columns": ["entity"],
        },
        "b": {
            "sheet": "B",
            "layout": "row_records",
            "title": "B",
            "unit": "count",
            "frequency": "Monthly",
            "header_row": 2,
            "data_start_row": 3,
            "date_column": 1,
            "dimension_columns": {"entity": 2},
            "series_start_column": 3,
            "identity_columns": ["entity"],
        },
    }

    payload = parse_apra_xlsx(
        workbook,
        publication_id="TEST_PUBLICATION",
        title="Test APRA publication",
        frequency="Monthly",
        table_maps=table_maps,
        table_id="b",
    )

    assert {series["dimensions"]["table"]["code"] for series in payload["series"]} == {"b"}
    assert payload["observations"][0]["series_id"] == "TEST_PUBLICATION:b:entity_b:metric"


def test_parse_apra_xlsx_selected_table_requires_matching_sheet() -> None:
    workbook = _xlsx_bytes(
        {
            "A": [
                [None],
                ["Period", "Entity", "Metric"],
                [datetime(2024, 1, 31), "Entity A", 1.0],
            ],
        }
    )
    table_maps = {
        "missing": {
            "sheet": "Missing",
            "layout": "row_records",
            "title": "Missing",
            "unit": "count",
            "frequency": "Monthly",
            "header_row": 2,
            "data_start_row": 3,
            "date_column": 1,
            "dimension_columns": {"entity": 2},
            "series_start_column": 3,
            "identity_columns": ["entity"],
        }
    }

    with pytest.raises(ValueError, match="did not contain sheet 'Missing'"):
        parse_apra_xlsx(
            workbook,
            publication_id="TEST_PUBLICATION",
            title="Test APRA publication",
            frequency="Monthly",
            table_maps=table_maps,
            table_id="missing",
        )


def _phi_workbook_with_year_ended_block() -> bytes:
    """Mirror the live PHI membership T1 sheet: an annual block, then a quarterly block."""
    return _xlsx_bytes(
        {
            "T1": [
                [None, "Coverage of hospital treatment tables", None, None, None, None],
                ["Year ended \n30 June", None, None, None, "NSW", "Aust."],
                ["Jun", "2024", "Coverage ('000)", None, 3947.3, 12300.1],
                [None, None, "% Population", None, 0.46, 0.45],
                ["Jun", "2025", "Coverage ('000)", None, 3990.4, 12531.0],
                [None, None, "% Population", None, 0.46, 0.45],
                ["Quarter ended", None, None, None, "NSW", "Aust."],
                ["Jun", "2025", "Coverage ('000)", None, 3990.4, 12531.0],
                [None, None, "% Population", None, 0.46, 0.45],
                ["Sep", "2025", "Coverage ('000)", None, 4010.0, 12633.9],
                [None, None, "% Population", None, 0.46, 0.45],
            ]
        }
    )


def test_parse_apra_xlsx_period_rows_start_after_period_marker() -> None:
    payload = parse_apra_xlsx(
        _phi_workbook_with_year_ended_block(),
        publication_id="TEST_PUBLICATION",
        title="Test PHI membership publication",
        frequency="Quarterly",
        table_maps={
            "t1": {
                "sheet": "T1",
                "layout": "period_rows",
                "title": "Hospital treatment membership coverage",
                "frequency": "Quarterly",
                "header_row": 2,
                "data_start_row": 3,
                "period_marker": "Quarter ended",
                "month_column": 1,
                "year_column": 2,
                "metric_column": 3,
                "series_start_column": 5,
            }
        },
    )

    coverage_dates = [
        item["date"]
        for item in payload["observations"]
        if item["series_id"] == "TEST_PUBLICATION:t1:aust:coverage_000"
    ]

    assert coverage_dates == ["2025-06-30", "2025-09-30"]


def test_phi_membership_catalogue_map_skips_year_ended_block() -> None:
    payload = parse_apra_xlsx(
        _phi_workbook_with_year_ended_block(),
        publication_id="APRA_PHI_MEMBERSHIP",
        title=APRA_CATALOGUE["APRA_PHI_MEMBERSHIP"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_PHI_MEMBERSHIP"]["tables"],
        table_id="t1",
    )

    keys = [(item["date"], item["series_id"]) for item in payload["observations"]]

    assert len(keys) == len(set(keys)), "duplicate (date, series_id) observations"
    assert "2024-06-30" not in {date for date, _ in keys}


def test_parse_apra_xlsx_streams_each_sheet_once(monkeypatch: pytest.MonkeyPatch) -> None:
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    calls = {"count": 0}
    original = ReadOnlyWorksheet._cells_by_row

    def counting(self, *args, **kwargs):
        calls["count"] += 1
        return original(self, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "_cells_by_row", counting)

    dates = [datetime(2020 + year, 3, 31) for year in range(5)]
    rows: list[list[object]] = [
        ["Key statistics", None, None, None, None, None],
        ["", *dates],
        ["Key figures", None, None, None, None, None],
    ]
    rows.extend([f"Metric {index}", *[float(index)] * 5] for index in range(30))

    parse_apra_xlsx(
        _xlsx_bytes({"Key Stats": rows}),
        publication_id="TEST_PUBLICATION",
        title="Test matrix publication",
        frequency="Quarterly",
        table_maps={
            "key_stats": {
                "sheet": "Key Stats",
                "layout": "matrix",
                "title": "Key statistics",
                "frequency": "Quarterly",
                "date_row": 2,
                "date_start_column": 2,
                "data_start_row": 3,
                "label_column": 1,
            }
        },
    )

    assert calls["count"] <= 2, f"sheet re-read {calls['count']} times"


def test_parse_apra_xlsx_does_not_load_external_link_caches(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # APRA's PHI membership workbook carries ~29 MB of external-link cache XML.
    # openpyxl parses it into memory by default (482 MB peak on a 512 MB host),
    # so the parser must opt out; the cached cell values already live in the sheets.
    seen: dict[str, object] = {}
    original = apra_parser.load_workbook

    def recording(*args, **kwargs):
        seen.update(kwargs)
        return original(*args, **kwargs)

    monkeypatch.setattr(apra_parser, "load_workbook", recording)

    parse_apra_xlsx(
        _xlsx_bytes({"T1": [[None, "Coverage"], ["Quarter ended", "Aust."], ["Dec", 1.0]]}),
        publication_id="TEST_PUBLICATION",
        title="Test",
        frequency="Quarterly",
        table_maps={
            "t1": {
                "sheet": "T1",
                "layout": "period_rows",
                "title": "T1",
                "frequency": "Quarterly",
                "header_row": 2,
                "data_start_row": 3,
                "month_column": 1,
                "year_column": 2,
                "metric_column": 3,
                "series_start_column": 2,
            }
        },
    )

    assert seen.get("read_only") is True
    assert seen.get("keep_links") is False


def _xlsx_bytes_with_formats(
    rows_by_sheet: dict[str, list[list[object]]],
    number_formats: dict[tuple[str, int, int], str],
) -> bytes:
    """Like _xlsx_bytes, but applies Excel number formats keyed by (sheet, row, column)."""
    workbook = Workbook()
    first = True
    sheets = {}
    for title, rows in rows_by_sheet.items():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = title
        for row in rows:
            sheet.append(row)
        sheets[title] = sheet
    for (title, row, column), number_format in number_formats.items():
        sheets[title].cell(row, column).number_format = number_format
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _key_stats_table_map() -> dict:
    return {
        "key_stats": {
            "sheet": "Key Stats",
            "layout": "matrix",
            "title": "Key statistics",
            "unit": None,
            "frequency": "Quarterly",
            "date_row": 2,
            "date_start_column": 2,
            "data_start_row": 3,
            "label_column": 1,
        }
    }


def test_parse_apra_xlsx_matrix_percent_formatted_rows_become_per_cent() -> None:
    workbook = _xlsx_bytes_with_formats(
        {
            "Key Stats": [
                ["Key figures", None, None],
                ["", datetime(2025, 12, 31), datetime(2026, 3, 31)],
                ["Key figures", None, None],
                ["Total capital base ($m)", 460000.0, 472998.4],
                ["Total capital ratio", 0.203, 0.203],
            ]
        },
        {("Key Stats", 5, 2): "0.0%", ("Key Stats", 5, 3): "0.0%"},
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="ADI_QUARTERLY_PERFORMANCE",
        title="Test",
        frequency="Quarterly",
        table_maps=_key_stats_table_map(),
    )

    series = {item["series_id"]: item for item in payload["series"]}
    ratio = series["ADI_QUARTERLY_PERFORMANCE:key_stats:key_figures:total_capital_ratio"]
    assert ratio["unit"] == "Per cent"
    assert ratio["decimals"] == 1
    ratio_values = [
        item["value"] for item in payload["observations"] if item["series_id"] == ratio["series_id"]
    ]
    assert ratio_values == [20.3, 20.3]


def test_parse_apra_xlsx_matrix_dollar_million_label_sets_unit_when_table_has_none() -> None:
    workbook = _xlsx_bytes_with_formats(
        {
            "Key Stats": [
                ["Key figures", None, None],
                ["", datetime(2025, 12, 31), datetime(2026, 3, 31)],
                ["Key figures", None, None],
                ["Total capital base ($m)", 460000.0, 472998.4],
            ]
        },
        {},
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="ADI_QUARTERLY_PERFORMANCE",
        title="Test",
        frequency="Quarterly",
        table_maps=_key_stats_table_map(),
    )

    assert payload["series"][0]["unit"] == "$ million"
    assert payload["observations"][1]["value"] == 472998.4


def test_parse_apra_xlsx_period_rows_percent_formatted_metric_overrides_table_unit() -> None:
    workbook = _xlsx_bytes_with_formats(
        {
            "T1": [
                [None, "Coverage of hospital treatment tables", None, None, None, None],
                ["Quarter ended", None, None, None, "NSW", "Aust."],
                ["Jun", "2026", "Coverage ('000)", None, 4100.0, 12823.752],
                [None, None, "% Population", None, 0.462, 0.4583],
            ]
        },
        {("T1", 4, 5): "0.0%", ("T1", 4, 6): "0.0%"},
    )

    payload = parse_apra_xlsx(
        workbook,
        publication_id="APRA_PHI_MEMBERSHIP",
        title=APRA_CATALOGUE["APRA_PHI_MEMBERSHIP"]["name"],
        frequency="Quarterly",
        table_maps=APRA_CATALOGUE["APRA_PHI_MEMBERSHIP"]["tables"],
        table_id="t1",
    )

    series = {item["series_id"]: item for item in payload["series"]}
    assert series["APRA_PHI_MEMBERSHIP:t1:aust:coverage_000"]["unit"] == "000 persons"
    assert series["APRA_PHI_MEMBERSHIP:t1:aust:population"]["unit"] == "Per cent"
    population = [
        item["value"]
        for item in payload["observations"]
        if item["series_id"] == "APRA_PHI_MEMBERSHIP:t1:aust:population"
    ]
    assert population == [45.83]
