from __future__ import annotations

import calendar
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from ausecon_mcp.errors import AuseconValidationError
from ausecon_mcp.models import Observation, SeriesDescriptor, parse_float

_SLUG_RE = re.compile(r"[^a-z0-9]+")
MAX_APRA_XLSX_BYTES = 50 * 1024 * 1024
MAX_APRA_XLSX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_APRA_XLSX_MEMBERS = 5000
MAX_APRA_TABLE_ROWS = 250_000
MAX_APRA_TABLE_COLUMNS = 512


def parse_apra_xlsx(
    content: bytes,
    *,
    publication_id: str,
    title: str,
    frequency: str,
    table_maps: dict[str, dict[str, Any]],
    table_id: str | None = None,
) -> dict[str, Any]:
    """Parse a curated APRA XLSX workbook into the normal retrieval response shape."""
    selected_maps = _select_table_maps(table_maps, table_id)
    _validate_xlsx_container(content)
    # keep_links=False: APRA workbooks can carry tens of MB of external-link cache XML
    # that openpyxl would otherwise parse into memory; cached values live in the sheets.
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True, keep_links=False)
    series_by_id: dict[str, dict[str, Any]] = {}
    observations: list[dict[str, Any]] = []

    for selected_table_id, table_map in selected_maps.items():
        sheet_name = str(table_map["sheet"])
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"APRA workbook did not contain sheet {sheet_name!r}.")
        sheet = workbook[sheet_name]
        _validate_sheet_bounds(sheet, selected_table_id)
        grid = _load_grid(sheet)
        layout = table_map["layout"]
        if layout == "row_records":
            table_series, table_observations = _parse_row_records(
                grid,
                publication_id=publication_id,
                table_id=selected_table_id,
                table_map=table_map,
            )
        elif layout == "matrix":
            table_series, table_observations = _parse_matrix(
                grid,
                publication_id=publication_id,
                table_id=selected_table_id,
                table_map=table_map,
            )
        elif layout == "period_rows":
            table_series, table_observations = _parse_period_rows(
                grid,
                publication_id=publication_id,
                table_id=selected_table_id,
                table_map=table_map,
            )
        else:
            raise ValueError(f"Unsupported APRA table layout: {layout!r}.")

        for item in table_series:
            series_by_id.setdefault(item["series_id"], item)
        observations.extend(table_observations)

    series_order = {series_id: index for index, series_id in enumerate(series_by_id)}
    return {
        "metadata": {
            "source": "apra",
            "dataset_id": publication_id,
            "frequency": frequency,
            "title": title,
        },
        "series": list(series_by_id.values()),
        "observations": sorted(
            observations,
            key=lambda item: (
                str(item["date"]),
                series_order.get(str(item["series_id"]), 0),
                str(item.get("dimensions", {})),
            ),
        ),
    }


def _select_table_maps(
    table_maps: dict[str, dict[str, Any]],
    table_id: str | None,
) -> dict[str, dict[str, Any]]:
    if table_id is None:
        return table_maps
    if table_id not in table_maps:
        known = ", ".join(sorted(table_maps))
        raise AuseconValidationError(
            f"Unknown APRA table {table_id!r}. Known tables: {known or '(none)'}."
        )
    return {table_id: table_maps[table_id]}


def _validate_xlsx_container(content: bytes) -> None:
    if len(content) > MAX_APRA_XLSX_BYTES:
        raise ValueError(
            "APRA workbook exceeds maximum APRA XLSX size "
            f"({len(content)} > {MAX_APRA_XLSX_BYTES} bytes)."
        )
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_APRA_XLSX_MEMBERS:
                raise ValueError(
                    "APRA workbook exceeds maximum XLSX member count "
                    f"({len(members)} > {MAX_APRA_XLSX_MEMBERS})."
                )
            uncompressed_bytes = sum(member.file_size for member in members)
    except zipfile.BadZipFile as exc:
        raise ValueError("APRA workbook is not a valid XLSX zip container.") from exc
    if uncompressed_bytes > MAX_APRA_XLSX_UNCOMPRESSED_BYTES:
        raise ValueError(
            "APRA workbook exceeds maximum uncompressed APRA XLSX payload "
            f"({uncompressed_bytes} > {MAX_APRA_XLSX_UNCOMPRESSED_BYTES} bytes)."
        )


def _validate_sheet_bounds(sheet: Any, table_id: str) -> None:
    max_row = int(sheet.max_row or 0)
    max_column = int(sheet.max_column or 0)
    if max_row > MAX_APRA_TABLE_ROWS:
        raise ValueError(
            f"APRA table {table_id!r} exceeds maximum APRA table rows "
            f"({max_row} > {MAX_APRA_TABLE_ROWS})."
        )
    if max_column > MAX_APRA_TABLE_COLUMNS:
        raise ValueError(
            f"APRA table {table_id!r} exceeds maximum APRA table columns "
            f"({max_column} > {MAX_APRA_TABLE_COLUMNS})."
        )


class Grid:
    """A read-only worksheet streamed exactly once into value and number-format grids.

    ``ReadOnlyWorksheet.cell()`` re-parses the sheet XML from the top on every call,
    which made large APRA workbooks quadratic to parse. Number formats are kept because
    APRA stores ratios as fractions displayed with an Excel percent format.
    """

    def __init__(self, sheet: Any) -> None:
        values: list[tuple[Any, ...]] = []
        formats: list[tuple[str | None, ...]] = []
        for row in sheet.iter_rows():
            values.append(tuple(cell.value for cell in row))
            formats.append(tuple(getattr(cell, "number_format", None) for cell in row))
        self._values = values
        self._formats = formats
        self.max_row = len(values)
        self.max_column = max((len(row) for row in values), default=0)

    def value(self, row: int, column: int) -> Any:
        return _lookup(self._values, row, column)

    def number_format(self, row: int, column: int) -> str | None:
        return _lookup(self._formats, row, column)


def _lookup(rows: list[tuple[Any, ...]], row: int, column: int) -> Any:
    if row < 1 or row > len(rows):
        return None
    values = rows[row - 1]
    if column < 1 or column > len(values):
        return None
    return values[column - 1]


def _load_grid(sheet: Any) -> Grid:
    return Grid(sheet)


def _grid_max_row(grid: Grid) -> int:
    return grid.max_row


def _grid_max_column(grid: Grid) -> int:
    return grid.max_column


def _cell_value(grid: Grid, row: int, column: int) -> Any:
    return grid.value(row, column)


def _cell_format(grid: Grid, row: int, column: int) -> str | None:
    return grid.number_format(row, column)


_PERCENT_DECIMALS_RE = re.compile(r"\.(0+)\s*%")
_DOLLAR_MILLION_LABEL_RE = re.compile(r"\(\$m\)\s*$", re.IGNORECASE)


def _is_percent_format(number_format: str | None) -> bool:
    return bool(number_format) and "%" in str(number_format)


def _apply_number_format(value: float | None, number_format: str | None) -> float | None:
    """Return the value as displayed: percent-formatted fractions become percentages."""
    if value is None or not _is_percent_format(number_format):
        return value
    return float(Decimal(str(value)) * 100)


def _series_unit(table_unit: Any, label: str, number_format: str | None) -> str | None:
    if _is_percent_format(number_format):
        return "Per cent"
    if table_unit is None and _DOLLAR_MILLION_LABEL_RE.search(label):
        return "$ million"
    return table_unit


def _series_decimals(number_format: str | None) -> int | None:
    if not _is_percent_format(number_format):
        return None
    match = _PERCENT_DECIMALS_RE.search(str(number_format))
    return len(match.group(1)) if match else 0


def _parse_row_records(
    grid: Grid,
    *,
    publication_id: str,
    table_id: str,
    table_map: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    header_row = int(table_map["header_row"])
    data_start_row = int(table_map["data_start_row"])
    date_column = int(table_map["date_column"])
    dimension_columns: dict[str, int] = dict(table_map.get("dimension_columns", {}))
    identity_columns: list[str] = list(table_map.get("identity_columns", []))
    series_start_column = int(table_map["series_start_column"])
    max_column = _grid_max_column(grid) or series_start_column

    metric_headers = {
        column: _clean_label(_cell_value(grid, header_row, column))
        for column in range(series_start_column, max_column + 1)
    }
    metric_headers = {column: label for column, label in metric_headers.items() if label}

    series_by_id: dict[str, dict[str, Any]] = {}
    observations: list[dict[str, Any]] = []

    for row_index in range(data_start_row, (_grid_max_row(grid) or data_start_row) + 1):
        parsed_date = _parse_date(_cell_value(grid, row_index, date_column))
        if parsed_date is None:
            continue
        dimension_values = {
            name: _clean_label(_cell_value(grid, row_index, column))
            for name, column in dimension_columns.items()
        }
        dimension_values = {name: value for name, value in dimension_values.items() if value}
        identity = _identity_slug(dimension_values, identity_columns)
        if not identity:
            identity = f"row_{row_index}"

        for column, metric_label in metric_headers.items():
            value, raw_value = _parse_observation_value(_cell_value(grid, row_index, column))
            if value is None and raw_value is None:
                continue
            number_format = _cell_format(grid, row_index, column)
            value = _apply_number_format(value, number_format)
            metric_slug = _slug(metric_label)
            series_id = f"{publication_id}:{table_id}:{identity}:{metric_slug}"
            dimensions = _row_dimensions(table_id, table_map, dimension_values)
            if series_id not in series_by_id:
                entity_label = _entity_label(dimension_values)
                label = f"{entity_label} - {metric_label}" if entity_label else metric_label
                series_by_id[series_id] = SeriesDescriptor(
                    series_id=series_id,
                    label=label,
                    unit=_series_unit(table_map.get("unit"), metric_label, number_format),
                    frequency=table_map.get("frequency"),
                    dimensions=dimensions,
                    source_key=metric_label,
                    decimals=_series_decimals(number_format),
                ).to_dict()
            observations.append(
                Observation(
                    date=parsed_date,
                    series_id=series_id,
                    value=value,
                    raw_value=raw_value,
                    dimensions=dimensions,
                ).to_dict()
            )

    return list(series_by_id.values()), observations


def _parse_matrix(
    grid: Grid,
    *,
    publication_id: str,
    table_id: str,
    table_map: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    date_row = int(table_map["date_row"])
    date_start_column = int(table_map["date_start_column"])
    data_start_row = int(table_map["data_start_row"])
    label_column = int(table_map["label_column"])

    date_columns: list[tuple[int, str]] = []
    for column in range(date_start_column, (_grid_max_column(grid) or date_start_column) + 1):
        parsed_date = _parse_date(_cell_value(grid, date_row, column))
        if parsed_date is not None:
            date_columns.append((column, parsed_date))
    if not date_columns:
        raise ValueError(f"APRA matrix table {table_id!r} did not contain date columns.")

    series_by_id: dict[str, dict[str, Any]] = {}
    observations: list[dict[str, Any]] = []
    section_label: str | None = None

    for row_index in range(data_start_row, (_grid_max_row(grid) or data_start_row) + 1):
        row_label = _clean_label(_cell_value(grid, row_index, label_column))
        if not row_label:
            continue
        values = [
            _parse_observation_value(_cell_value(grid, row_index, column))
            for column, _ in date_columns
        ]
        formats = [_cell_format(grid, row_index, column) for column, _ in date_columns]
        has_observations = any(value is not None or raw is not None for value, raw in values)
        if not has_observations:
            if not row_label.lower().endswith(":"):
                section_label = row_label
            continue

        metric_slug = _slug(row_label)
        section_slug = _slug(section_label) if section_label else None
        series_id = (
            f"{publication_id}:{table_id}:{section_slug}:{metric_slug}"
            if section_slug
            else f"{publication_id}:{table_id}:{metric_slug}"
        )
        dimensions = _matrix_dimensions(table_id, table_map, section_label)
        if series_id not in series_by_id:
            label = f"{section_label}: {row_label}" if section_label else row_label
            number_format = next(
                (
                    fmt
                    for (value, _raw), fmt in zip(values, formats, strict=True)
                    if value is not None
                ),
                None,
            )
            series_by_id[series_id] = SeriesDescriptor(
                series_id=series_id,
                label=label,
                unit=_series_unit(table_map.get("unit"), row_label, number_format),
                frequency=table_map.get("frequency"),
                dimensions=dimensions,
                source_key=row_label,
                decimals=_series_decimals(number_format),
            ).to_dict()

        for (_column, parsed_date), (value, raw_value), number_format in zip(
            date_columns, values, formats, strict=True
        ):
            if value is None and raw_value is None:
                continue
            observations.append(
                Observation(
                    date=parsed_date,
                    series_id=series_id,
                    value=_apply_number_format(value, number_format),
                    raw_value=raw_value,
                    dimensions=dimensions,
                ).to_dict()
            )

    return list(series_by_id.values()), observations


def _parse_period_rows(
    grid: Grid,
    *,
    publication_id: str,
    table_id: str,
    table_map: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    header_row = int(table_map["header_row"])
    data_start_row = int(table_map["data_start_row"])
    month_column = int(table_map["month_column"])
    year_column = int(table_map["year_column"])
    metric_column = int(table_map["metric_column"])
    series_start_column = int(table_map["series_start_column"])

    region_headers = {
        column: _clean_label(_cell_value(grid, header_row, column))
        for column in range(
            series_start_column, (_grid_max_column(grid) or series_start_column) + 1
        )
    }
    region_headers = {column: label for column, label in region_headers.items() if label}
    if not region_headers:
        raise ValueError(f"APRA period-row table {table_id!r} did not contain region columns.")

    period_marker = _clean_label(table_map.get("period_marker"))
    if period_marker:
        marker = period_marker.lower()
        for row_index in range(header_row, _grid_max_row(grid) + 1):
            text = _clean_label(_cell_value(grid, row_index, month_column))
            if text and text.lower().startswith(marker):
                data_start_row = row_index + 1
                break

    series_by_id: dict[str, dict[str, Any]] = {}
    observations: list[dict[str, Any]] = []
    current_month: str | None = None
    current_year: str | None = None

    for row_index in range(data_start_row, (_grid_max_row(grid) or data_start_row) + 1):
        month_text = _clean_label(_cell_value(grid, row_index, month_column))
        year_text = _clean_label(_cell_value(grid, row_index, year_column))
        if month_text:
            current_month = month_text
        if year_text:
            current_year = year_text

        parsed_date = _parse_period_row_date(current_month, current_year)
        if parsed_date is None:
            continue
        metric_label = _clean_label(_cell_value(grid, row_index, metric_column))
        if not metric_label:
            continue

        metric_slug = _slug(metric_label)
        for column, region_label in region_headers.items():
            value, raw_value = _parse_observation_value(_cell_value(grid, row_index, column))
            if value is None and raw_value is None:
                continue
            number_format = _cell_format(grid, row_index, column)
            value = _apply_number_format(value, number_format)
            region_slug = _slug(region_label)
            series_id = f"{publication_id}:{table_id}:{region_slug}:{metric_slug}"
            dimensions = _period_row_dimensions(table_id, table_map, region_label, metric_label)
            if series_id not in series_by_id:
                series_by_id[series_id] = SeriesDescriptor(
                    series_id=series_id,
                    label=f"{region_label} - {metric_label}",
                    unit=_series_unit(table_map.get("unit"), metric_label, number_format),
                    frequency=table_map.get("frequency"),
                    dimensions=dimensions,
                    source_key=metric_label,
                    decimals=_series_decimals(number_format),
                ).to_dict()
            observations.append(
                Observation(
                    date=parsed_date,
                    series_id=series_id,
                    value=value,
                    raw_value=raw_value,
                    dimensions=dimensions,
                ).to_dict()
            )

    return list(series_by_id.values()), observations


def _row_dimensions(
    table_id: str,
    table_map: dict[str, Any],
    dimension_values: dict[str, str],
) -> dict[str, dict[str, str]]:
    dimensions = {
        "table": {
            "code": table_id,
            "label": str(table_map.get("title") or table_id),
        }
    }
    for name, value in dimension_values.items():
        dimensions[name] = {"code": value, "label": value}
    return dimensions


def _matrix_dimensions(
    table_id: str,
    table_map: dict[str, Any],
    section_label: str | None,
) -> dict[str, dict[str, str]]:
    dimensions = {
        "table": {
            "code": table_id,
            "label": str(table_map.get("title") or table_id),
        }
    }
    if section_label:
        dimensions["section"] = {"code": _slug(section_label), "label": section_label}
    return dimensions


def _period_row_dimensions(
    table_id: str,
    table_map: dict[str, Any],
    region_label: str,
    metric_label: str,
) -> dict[str, dict[str, str]]:
    return {
        "table": {
            "code": table_id,
            "label": str(table_map.get("title") or table_id),
        },
        "region": {"code": _slug(region_label), "label": region_label},
        "metric": {"code": _slug(metric_label), "label": metric_label},
    }


def _parse_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%b %Y", "%B %Y"):
            try:
                parsed = datetime.strptime(text, fmt)
            except ValueError:
                continue
            day = calendar.monthrange(parsed.year, parsed.month)[1]
            return date(parsed.year, parsed.month, day).isoformat()
        try:
            return date.fromisoformat(text).isoformat()
        except ValueError:
            return None
    return None


def _parse_period_row_date(month_text: str | None, year_text: str | None) -> str | None:
    if not month_text or not year_text:
        return None
    try:
        year = int(float(year_text))
    except ValueError:
        return None
    for fmt in ("%b", "%B"):
        try:
            parsed = datetime.strptime(month_text[:3] if fmt == "%b" else month_text, fmt)
        except ValueError:
            continue
        month = parsed.month
        day = calendar.monthrange(year, month)[1]
        return date(year, month, day).isoformat()
    return None


def _parse_observation_value(value: Any) -> tuple[float | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, bool):
        return None, str(value)
    if isinstance(value, int | float):
        return float(value), None
    text = str(value).strip()
    if not text:
        return None, None
    try:
        return parse_float(text), None
    except ValueError:
        return None, text


def _identity_slug(
    dimension_values: dict[str, str],
    identity_columns: list[str],
) -> str:
    values = [dimension_values.get(name, "") for name in identity_columns]
    return ":".join(_slug(value) for value in values if value)


def _entity_label(dimension_values: dict[str, str]) -> str | None:
    for key in ("institution", "entity", "abn"):
        value = dimension_values.get(key)
        if value:
            return value
    return None


def _clean_label(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _slug(value: str) -> str:
    slug = _SLUG_RE.sub("_", value.strip().lower()).strip("_")
    return slug or "unknown"
