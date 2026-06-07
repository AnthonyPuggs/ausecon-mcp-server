from __future__ import annotations

import argparse
import calendar
import json
from datetime import datetime
from io import BytesIO
from statistics import mean
from time import perf_counter
from typing import Any

from openpyxl import Workbook, load_workbook

from ausecon_mcp.parsers.apra_xlsx import parse_apra_xlsx

PUBLICATION_ID = "BENCHMARK_APRA"


def _xlsx_bytes(sheet_name: str, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _workbook_dimensions(content: bytes) -> tuple[int, int]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return int(sheet.max_row or 0), int(sheet.max_column or 0)


def _month_end(month_index: int) -> datetime:
    month = month_index % 12 + 1
    year = 2024 + month_index // 12
    day = calendar.monthrange(year, month)[1]
    return datetime(year, month, day)


def build_row_records_case(row_count: int) -> tuple[bytes, dict[str, dict[str, Any]]]:
    rows: list[list[object]] = [
        ["($ million)"],
        ["Period", "ABN", "Institution Name", "Total assets", "Deposits"],
    ]
    for index in range(row_count):
        rows.append(
            [
                _month_end(index),
                10_000_000_000 + index,
                f"Benchmark ADI {index}",
                float(index * 10 + 1),
                float(index * 10 + 2),
            ]
        )
    return _xlsx_bytes("Row Records", rows), {
        "row_records": {
            "sheet": "Row Records",
            "layout": "row_records",
            "title": "Benchmark row records",
            "unit": "$ million",
            "frequency": "Monthly",
            "header_row": 2,
            "data_start_row": 3,
            "date_column": 1,
            "dimension_columns": {"abn": 2, "institution": 3},
            "series_start_column": 4,
            "identity_columns": ["abn"],
        }
    }


def build_matrix_case(row_count: int, date_count: int) -> tuple[bytes, dict[str, dict[str, Any]]]:
    dates = [_month_end(index) for index in range(date_count)]
    rows: list[list[object]] = [[None, *dates]]
    rows.append(["Financial position", *([None] * date_count)])
    for row_index in range(row_count):
        rows.append(
            [
                f"Metric {row_index}",
                *[float(row_index * 100 + date_index) for date_index in range(date_count)],
            ]
        )
    return _xlsx_bytes("Matrix", rows), {
        "matrix": {
            "sheet": "Matrix",
            "layout": "matrix",
            "title": "Benchmark matrix",
            "unit": "$ million",
            "frequency": "Quarterly",
            "date_row": 1,
            "date_start_column": 2,
            "data_start_row": 2,
            "label_column": 1,
        }
    }


def build_period_rows_case(row_count: int) -> tuple[bytes, dict[str, dict[str, Any]]]:
    rows: list[list[object]] = [["Month", "Year", "Metric", "NSW", "VIC"]]
    for index in range(row_count):
        month = _month_end(index).strftime("%B")
        year = str(_month_end(index).year)
        rows.append([month, year, f"Metric {index}", float(index + 1), float(index + 2)])
    return _xlsx_bytes("Period Rows", rows), {
        "period_rows": {
            "sheet": "Period Rows",
            "layout": "period_rows",
            "title": "Benchmark period rows",
            "unit": "$ million",
            "frequency": "Monthly",
            "header_row": 1,
            "data_start_row": 2,
            "month_column": 1,
            "year_column": 2,
            "metric_column": 3,
            "series_start_column": 4,
        }
    }


def _benchmark_case(
    layout: str,
    content: bytes,
    table_maps: dict[str, dict[str, Any]],
    *,
    repeats: int,
) -> dict[str, Any]:
    timings: list[float] = []
    payload: dict[str, Any] | None = None
    for _ in range(repeats):
        start = perf_counter()
        payload = parse_apra_xlsx(
            content,
            publication_id=PUBLICATION_ID,
            title=f"Benchmark {layout}",
            frequency=str(next(iter(table_maps.values())).get("frequency", "Monthly")),
            table_maps=table_maps,
        )
        timings.append((perf_counter() - start) * 1000)

    if payload is None:
        raise RuntimeError("Benchmark did not parse any payload.")
    elapsed_ms = mean(timings)
    observations = len(payload["observations"])
    observations_per_second = observations / (elapsed_ms / 1000) if elapsed_ms else 0.0
    rows, columns = _workbook_dimensions(content)

    return {
        "layout": layout,
        "rows": rows,
        "columns": columns,
        "workbook_bytes": len(content),
        "series": len(payload["series"]),
        "observations": observations,
        "elapsed_ms": round(elapsed_ms, 3),
        "observations_per_second": round(observations_per_second, 3),
    }


def run_benchmark(row_count: int, date_count: int, repeats: int) -> dict[str, Any]:
    cases = [
        ("row_records", *build_row_records_case(row_count)),
        ("matrix", *build_matrix_case(row_count, date_count)),
        ("period_rows", *build_period_rows_case(row_count)),
    ]
    return {
        "benchmark": "apra_parser",
        "row_count": row_count,
        "date_count": date_count,
        "repeats": repeats,
        "results": [
            _benchmark_case(layout, content, table_maps, repeats=repeats)
            for layout, content, table_maps in cases
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Benchmark synthetic APRA XLSX parser layouts.")
    parser.add_argument("--row-count", type=int, default=1000)
    parser.add_argument("--date-count", type=int, default=24)
    parser.add_argument("--repeats", type=int, default=3)
    args = parser.parse_args()

    if args.row_count < 1:
        parser.error("--row-count must be positive")
    if args.date_count < 1:
        parser.error("--date-count must be positive")
    if args.repeats < 1:
        parser.error("--repeats must be positive")

    print(json.dumps(run_benchmark(args.row_count, args.date_count, args.repeats), indent=2))


if __name__ == "__main__":
    main()
