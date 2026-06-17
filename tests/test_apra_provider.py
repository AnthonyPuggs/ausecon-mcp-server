from __future__ import annotations

import asyncio
import json
import time
from datetime import datetime
from io import BytesIO

import pytest
import respx
from httpx import ConnectTimeout, Response
from openpyxl import Workbook

from ausecon_mcp.cache import TTLCache
from ausecon_mcp.errors import AuseconParseError
from ausecon_mcp.providers.apra import APRAProvider, resolve_apra_download_url


def _xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table 1"
    sheet.append(["($million)"])
    sheet.append(["Period", "ABN", "Institution Name", "Total residents assets"])
    sheet.append([datetime(2024, 1, 31), 11111111111, "Example Bank", 100.5])
    sheet.append([datetime(2024, 2, 29), 11111111111, "Example Bank", 110.0])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _two_table_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table 1"
    sheet.append(["($million)"])
    sheet.append(["Period", "ABN", "Institution Name", "Total residents assets"])
    sheet.append([datetime(2024, 1, 31), 11111111111, "Example Bank", 100.5])

    second = workbook.create_sheet("Table 2")
    second.append(["($million)"])
    second.append(["Period", "ABN", "Institution Name", "Other assets"])
    second.append([datetime(2024, 1, 31), 22222222222, "Second Bank", 200.0])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _catalogue() -> dict:
    return {
        "TEST_PUBLICATION": {
            "id": "TEST_PUBLICATION",
            "name": "Test APRA publication",
            "description": "Selected APRA test data.",
            "category": "banking",
            "frequency": "Monthly",
            "frequencies": ["M"],
            "geographies": ["aus"],
            "tags": ["apra", "adi"],
            "landing_url": "https://www.apra.gov.au/test-statistics",
            "link_patterns": [r"back-series.*xlsx"],
            "tables": {
                "table_1": {
                    "sheet": "Table 1",
                    "layout": "row_records",
                    "title": "Table 1",
                    "unit": "$ million",
                    "frequency": "Monthly",
                    "header_row": 2,
                    "data_start_row": 3,
                    "date_column": 1,
                    "dimension_columns": {"abn": 2, "institution": 3},
                    "series_start_column": 4,
                    "identity_columns": ["abn"],
                },
                "table_2": {
                    "sheet": "Table 2",
                    "layout": "row_records",
                    "title": "Table 2",
                    "unit": "$ million",
                    "frequency": "Monthly",
                    "header_row": 2,
                    "data_start_row": 3,
                    "date_column": 1,
                    "dimension_columns": {"abn": 2, "institution": 3},
                    "series_start_column": 4,
                    "identity_columns": ["abn"],
                },
            },
        }
    }


def _seed_catalogue() -> dict:
    catalogue = _catalogue()
    catalogue["TEST_PUBLICATION"]["url_seeds"] = [
        {
            "url": "https://www.apra.gov.au/sites/default/files/seeded-test.xlsx",
            "label": "Test back-series March 2026 XLSX",
            "checked_at": "2026-05-20T00:00:00Z",
        }
    ]
    catalogue["TEST_PUBLICATION"]["framework_breaks"] = [
        {
            "date": "2023-07-01",
            "label": "AASB 17 transition",
            "description": "Insurance accounting framework changed from this date.",
        }
    ]
    return catalogue


@pytest.mark.parametrize(
    "href",
    [
        "https://evil.example/apra-back-series.xlsx",
        "//evil.example/apra-back-series.xlsx",
        "http://www.apra.gov.au/sites/default/files/apra-back-series.xlsx",
    ],
)
def test_resolve_apra_download_url_rejects_non_apra_or_non_https_xlsx_links(
    href: str,
) -> None:
    html = f'<a href="{href}">Test back-series March 2026 XLSX</a>'

    with pytest.raises(AuseconParseError, match="trusted APRA HTTPS"):
        resolve_apra_download_url(
            html,
            base_url="https://www.apra.gov.au/test-statistics",
            patterns=[r"back-series.*xlsx"],
        )


@pytest.mark.asyncio
async def test_apra_provider_resolves_landing_page_xlsx_and_uses_cache() -> None:
    provider = APRAProvider(catalogue=_catalogue())
    html = '<a href="/sites/default/files/test.xlsx">Test back-series March 2026 XLSX</a>'

    with respx.mock(assert_all_called=True) as router:
        landing_route = router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        xlsx_route = router.get("https://www.apra.gov.au/sites/default/files/test.xlsx").mock(
            return_value=Response(200, content=_xlsx_bytes())
        )

        first = await provider.get_data("TEST_PUBLICATION", table_id="table_1", last_n=1)
        second = await provider.get_data("TEST_PUBLICATION", table_id="table_1", last_n=1)

    assert landing_route.call_count == 1
    assert xlsx_route.call_count == 1
    assert first == second
    assert first["metadata"]["source"] == "apra"
    assert first["metadata"]["dataset_id"] == "TEST_PUBLICATION"
    assert first["metadata"]["retrieval_url"].endswith("/sites/default/files/test.xlsx")
    assert first["metadata"]["truncated"] is True
    assert first["observations"][0]["date"] == "2024-02-29"


@pytest.mark.asyncio
async def test_apra_provider_falls_back_to_trusted_url_seed_when_landing_page_drifts() -> None:
    provider = APRAProvider(catalogue=_seed_catalogue())
    html = "<html><body><p>No matching workbook link today.</p></body></html>"

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        seeded_route = router.get(
            "https://www.apra.gov.au/sites/default/files/seeded-test.xlsx"
        ).mock(return_value=Response(200, content=_xlsx_bytes()))

        result = await provider.get_data("TEST_PUBLICATION", table_id="table_1")

    assert seeded_route.call_count == 1
    assert result["metadata"]["retrieval_url"].endswith("/seeded-test.xlsx")
    assert result["metadata"]["apra_url_resolution"] == {
        "strategy": "seed_manifest",
        "seed_checked_at": "2026-05-20T00:00:00Z",
    }


@pytest.mark.asyncio
async def test_apra_provider_uses_bundled_seed_before_catalogue_fallback(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    catalogue = _catalogue()
    catalogue["TEST_PUBLICATION"]["fallback_url"] = (
        "https://www.apra.gov.au/sites/default/files/catalogue-test.xlsx"
    )
    monkeypatch.setattr(
        "ausecon_mcp.providers.apra._bundled_url_seeds",
        lambda: {
            "TEST_PUBLICATION": [
                {
                    "url": "https://www.apra.gov.au/sites/default/files/bundled-seed.xlsx",
                    "label": "Test back-series March 2026 XLSX",
                    "checked_at": "2026-05-21T00:00:00Z",
                }
            ]
        },
    )
    provider = APRAProvider(catalogue=catalogue)
    html = "<html><body><p>No matching workbook link today.</p></body></html>"

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        seed_route = router.get(
            "https://www.apra.gov.au/sites/default/files/bundled-seed.xlsx"
        ).mock(return_value=Response(200, content=_xlsx_bytes()))

        result = await provider.get_data("TEST_PUBLICATION", table_id="table_1")

    assert seed_route.call_count == 1
    assert result["metadata"]["retrieval_url"].endswith("/bundled-seed.xlsx")
    assert result["metadata"]["apra_url_resolution"] == {
        "strategy": "seed_manifest",
        "seed_checked_at": "2026-05-21T00:00:00Z",
    }


@pytest.mark.asyncio
async def test_apra_provider_uses_catalogue_fallback_after_seed_manifest_miss(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    catalogue = _catalogue()
    catalogue["TEST_PUBLICATION"]["fallback_url"] = (
        "https://www.apra.gov.au/sites/default/files/catalogue-test.xlsx"
    )
    monkeypatch.setattr("ausecon_mcp.providers.apra._bundled_url_seeds", lambda: {})
    provider = APRAProvider(catalogue=catalogue)
    html = "<html><body><p>No matching workbook link today.</p></body></html>"

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        fallback_route = router.get(
            "https://www.apra.gov.au/sites/default/files/catalogue-test.xlsx"
        ).mock(return_value=Response(200, content=_xlsx_bytes()))

        result = await provider.get_data("TEST_PUBLICATION", table_id="table_1")

    assert fallback_route.call_count == 1
    assert result["metadata"]["retrieval_url"].endswith("/catalogue-test.xlsx")
    assert result["metadata"]["apra_url_resolution"] == {
        "strategy": "catalogue_fallback",
        "seed_checked_at": None,
    }


@pytest.mark.asyncio
async def test_apra_provider_rejects_untrusted_seed_urls() -> None:
    catalogue = _catalogue()
    catalogue["TEST_PUBLICATION"]["url_seeds"] = [
        {
            "url": "https://evil.example/seeded-test.xlsx",
            "label": "Test back-series March 2026 XLSX",
            "checked_at": "2026-05-20T00:00:00Z",
        }
    ]
    provider = APRAProvider(catalogue=catalogue)
    html = "<html><body><p>No matching workbook link today.</p></body></html>"

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )

        with pytest.raises(AuseconParseError, match="trusted APRA HTTPS"):
            await provider.get_data("TEST_PUBLICATION", table_id="table_1")


@pytest.mark.asyncio
async def test_apra_provider_rejects_untrusted_catalogue_fallback_urls(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    catalogue = _catalogue()
    catalogue["TEST_PUBLICATION"]["fallback_url"] = "https://evil.example/catalogue-test.xlsx"
    monkeypatch.setattr("ausecon_mcp.providers.apra._bundled_url_seeds", lambda: {})
    provider = APRAProvider(catalogue=catalogue)
    html = "<html><body><p>No matching workbook link today.</p></body></html>"

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )

        with pytest.raises(AuseconParseError, match="trusted APRA HTTPS"):
            await provider.get_data("TEST_PUBLICATION", table_id="table_1")


@pytest.mark.asyncio
async def test_apra_provider_stamps_framework_break_warnings() -> None:
    provider = APRAProvider(catalogue=_seed_catalogue())
    html = '<a href="/sites/default/files/test.xlsx">Test back-series March 2026 XLSX</a>'

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        router.get("https://www.apra.gov.au/sites/default/files/test.xlsx").mock(
            return_value=Response(200, content=_xlsx_bytes())
        )

        result = await provider.get_data("TEST_PUBLICATION", table_id="table_1")

    assert result["metadata"]["framework_breaks"] == [
        {
            "date": "2023-07-01",
            "label": "AASB 17 transition",
            "description": "Insurance accounting framework changed from this date.",
        }
    ]
    assert result["metadata"]["warnings"] == [
        "AASB 17 transition on 2023-07-01: Insurance accounting framework changed from this date."
    ]


@pytest.mark.asyncio
async def test_apra_provider_passes_requested_table_id_to_parser(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    provider = APRAProvider(catalogue=_catalogue())
    html = '<a href="/sites/default/files/test.xlsx">Test back-series March 2026 XLSX</a>'
    calls: list[dict] = []

    def fake_parse_apra_xlsx(content: bytes, **kwargs) -> dict:
        calls.append(kwargs)
        return {
            "metadata": {
                "source": "apra",
                "dataset_id": kwargs["publication_id"],
                "frequency": kwargs["frequency"],
                "title": kwargs["title"],
            },
            "series": [],
            "observations": [],
        }

    monkeypatch.setattr("ausecon_mcp.providers.apra.parse_apra_xlsx", fake_parse_apra_xlsx)

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        router.get("https://www.apra.gov.au/sites/default/files/test.xlsx").mock(
            return_value=Response(200, content=b"not-used-by-stub")
        )

        await provider.get_data("TEST_PUBLICATION", table_id="table_1")

    assert calls
    assert calls[0]["table_id"] == "table_1"


@pytest.mark.asyncio
async def test_apra_provider_uses_table_aware_cache_entries() -> None:
    provider = APRAProvider(catalogue=_catalogue())
    html = '<a href="/sites/default/files/test.xlsx">Test back-series March 2026 XLSX</a>'

    with respx.mock(assert_all_called=True) as router:
        landing_route = router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        xlsx_route = router.get("https://www.apra.gov.au/sites/default/files/test.xlsx").mock(
            return_value=Response(200, content=_two_table_xlsx_bytes())
        )

        first = await provider.get_data("TEST_PUBLICATION", table_id="table_1")
        second = await provider.get_data("TEST_PUBLICATION", table_id="table_1")
        third = await provider.get_data("TEST_PUBLICATION", table_id="table_2")

    assert first == second
    assert landing_route.call_count == 2
    assert xlsx_route.call_count == 2
    assert {series["dimensions"]["table"]["code"] for series in first["series"]} == {"table_1"}
    assert {series["dimensions"]["table"]["code"] for series in third["series"]} == {"table_2"}


@pytest.mark.asyncio
async def test_apra_provider_filters_series_ids_and_dates_after_cache() -> None:
    provider = APRAProvider(catalogue=_catalogue())
    html = '<a href="/sites/default/files/test.xlsx">Test back-series March 2026 XLSX</a>'
    series_id = "TEST_PUBLICATION:table_1:11111111111:total_residents_assets"

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        router.get("https://www.apra.gov.au/sites/default/files/test.xlsx").mock(
            return_value=Response(200, content=_xlsx_bytes())
        )

        result = await provider.get_data(
            "TEST_PUBLICATION",
            table_id="table_1",
            series_ids=[series_id],
            start_date="2024-02-01",
        )

    assert [item["date"] for item in result["observations"]] == ["2024-02-29"]
    assert result["series"][0]["series_id"] == series_id


@pytest.mark.asyncio
async def test_apra_provider_returns_stale_payload_on_upstream_failure(_isolated_cache_dir) -> None:
    cache = TTLCache(ttl_seconds=60)
    provider = APRAProvider(cache=cache, catalogue=_catalogue())
    html = '<a href="/sites/default/files/test.xlsx">Test back-series March 2026 XLSX</a>'

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        router.get("https://www.apra.gov.au/sites/default/files/test.xlsx").mock(
            return_value=Response(200, content=_xlsx_bytes())
        )
        await provider.get_data("TEST_PUBLICATION", table_id="table_1")

    (file,) = _isolated_cache_dir.glob("*.json")
    data = json.loads(file.read_text())
    data["expires_at"] = 0.0
    file.write_text(json.dumps(data))

    fresh_provider = APRAProvider(cache=TTLCache(ttl_seconds=60), catalogue=_catalogue())

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            side_effect=ConnectTimeout("network down")
        )

        result = await fresh_provider.get_data("TEST_PUBLICATION", table_id="table_1")

    assert result["metadata"]["stale"] is True
    assert result["observations"]


@pytest.mark.asyncio
async def test_apra_provider_parses_workbooks_off_event_loop(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    provider = APRAProvider(catalogue=_catalogue())
    html = '<a href="/sites/default/files/test.xlsx">Test back-series March 2026 XLSX</a>'
    parse_started_at: list[float] = []

    def slow_parse_apra_xlsx(content: bytes, **kwargs) -> dict:
        parse_started_at.append(time.perf_counter())
        time.sleep(0.5)
        return {
            "metadata": {
                "source": "apra",
                "dataset_id": kwargs["publication_id"],
                "frequency": kwargs["frequency"],
                "title": kwargs["title"],
            },
            "series": [],
            "observations": [],
        }

    monkeypatch.setattr("ausecon_mcp.providers.apra.parse_apra_xlsx", slow_parse_apra_xlsx)

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        router.get("https://www.apra.gov.au/sites/default/files/test.xlsx").mock(
            return_value=Response(200, content=b"not-used-by-stub")
        )

        task = asyncio.create_task(provider.get_data("TEST_PUBLICATION", table_id="table_1"))
        wait_started = time.perf_counter()
        while not parse_started_at and time.perf_counter() - wait_started < 1.0:
            await asyncio.sleep(0.01)

        assert parse_started_at
        assert time.perf_counter() - parse_started_at[0] < 0.25
        await task


@pytest.mark.asyncio
async def test_apra_provider_rejects_untrusted_final_host_after_redirect() -> None:
    """Post-redirect host re-validation rejects a final host outside the trusted APRA host set."""
    provider = APRAProvider(catalogue=_catalogue())
    html = '<a href="/sites/default/files/test.xlsx">Test back-series March 2026 XLSX</a>'

    with respx.mock(assert_all_called=True) as router:
        router.get("https://www.apra.gov.au/test-statistics").mock(
            return_value=Response(200, text=html)
        )
        # The XLSX URL redirects to an untrusted host.
        router.get("https://www.apra.gov.au/sites/default/files/test.xlsx").mock(
            return_value=Response(301, headers={"Location": "https://evil.example/file.xlsx"})
        )
        router.get("https://evil.example/file.xlsx").mock(
            return_value=Response(200, content=_xlsx_bytes())
        )

        with pytest.raises(AuseconParseError, match="untrusted host"):
            await provider.get_data("TEST_PUBLICATION", table_id="table_1")


@pytest.mark.asyncio
async def test_apra_provider_rejects_unknown_publication() -> None:
    provider = APRAProvider(catalogue=_catalogue())

    with pytest.raises(ValueError, match="Unknown APRA publication"):
        await provider.get_data("UNKNOWN")
